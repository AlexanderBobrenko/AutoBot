# -*- coding: utf-8 -*-
from __future__ import annotations

import asyncio
import re
from datetime import date, datetime, timedelta
from io import BytesIO

from aiogram import Bot, Dispatcher, Router, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

from config import ADMIN_IDS, CURRENCIES, CURRENCY_SYMBOLS, EUR_RUB_RATE, TOKEN, USD_RUB_RATE
from database import Database, FuelDatabase
from keyboards import (
    get_currency_keyboard,
    get_categories_keyboard,
    get_main_keyboard,
    get_period_keyboard,
    get_pagination_keyboard,
    get_reminder_periodicity_keyboard,
    get_reminder_types_keyboard,
    get_search_menu_keyboard,
    MAIN_BUTTON_ADD,
    MAIN_BUTTON_ADD_FUEL,
    MAIN_BUTTON_TODAY,
    MAIN_BUTTON_DYNAMIC_EXPENSES,
    MAIN_BUTTON_EXPORT_EXCEL,
    MAIN_BUTTON_FUEL_STATS,
    MAIN_BUTTON_GRAPH_EXPENSES,
    MAIN_BUTTON_GRAPH_FUEL,
    MAIN_BUTTON_WEEK,
    MAIN_BUTTON_PROFILE,
    MAIN_BUTTON_REMINDERS,
    MAIN_BUTTON_REMINDERS_LIST,
)

DB = Database("expenses.sqlite3")
FUEL_DB = FuelDatabase("expenses.sqlite3")

CATEGORIES = ["Топливо", "Ремонт", "Мойка", "Страховка", "Другое"]
CATEGORY_DISPLAY = {
    "Топливо": "⛽ Топливо",
    "Ремонт": "🛠 Ремонт",
    "Мойка": "🚿 Мойка",
    "Страховка": "🧾 Страховка",
    "Другое": "📦 Другое",
}

FUEL_CAR_MODEL_DEFAULT = "default"

REMINDER_TYPES = ["ТО", "Страховка", "Налог", "Шиномонтаж", "Другое"]
REMINDER_PERIODICITIES = {"once": "разово", "monthly": "ежемесячно", "yearly": "ежегодно"}

SEARCH_PER_PAGE = 10


def _parse_float_ru(text: str) -> float | None:
    """
    Парсит числа вида "123.45" и "123,45".
    """
    cleaned = (text or "").strip().replace(" ", "")
    if not cleaned:
        return None
    cleaned = cleaned.replace(",", ".")
    try:
        value = float(cleaned)
    except ValueError:
        return None
    return value


def _parse_int_ru(text: str) -> int | None:
    cleaned = (text or "").strip().replace(" ", "")
    if not cleaned.isdigit():
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _format_amount(amount_rub: float | int, currency: str) -> str:
    amount = float(amount_rub)
    currency = currency.upper()
    symbol = CURRENCY_SYMBOLS.get(currency, "")
    if currency == "RUB":
        return f"{int(round(amount))}{symbol}"
    if currency == "USD":
        val = amount / USD_RUB_RATE
        return f"{val:.2f}{symbol}"
    if currency == "EUR":
        val = amount / EUR_RUB_RATE
        return f"{val:.2f}{symbol}"
    return f"{int(round(amount))}₽"


def _convert_amount_rub(amount_rub: float | int, currency: str) -> float:
    currency = currency.upper()
    amount = float(amount_rub)
    if currency == "RUB":
        return amount
    if currency == "USD":
        return amount / USD_RUB_RATE
    if currency == "EUR":
        return amount / EUR_RUB_RATE
    return amount


def _today_iso() -> str:
    return date.today().isoformat()


def _yesterday_iso() -> str:
    return (date.today() - timedelta(days=1)).isoformat()


def _this_week_range() -> tuple[str, str]:
    # последние 7 дней (включая сегодня)
    end = date.today()
    start = end - timedelta(days=6)
    return start.isoformat(), end.isoformat()


def _this_month_range() -> tuple[str, str]:
    # последние 30 дней (включая сегодня)
    end = date.today()
    start = end - timedelta(days=29)
    return start.isoformat(), end.isoformat()


async def ensure_user_not_banned_message(message: Message) -> bool:
    user_id = int(message.from_user.id)
    username = message.from_user.username
    DB.upsert_user(user_id=user_id, username=username)

    if user_id not in ADMIN_IDS and DB.is_user_banned(user_id):
        await message.answer("Доступ запрещен. Вы заблокированы администратором.")
        return False
    return True


async def ensure_user_not_banned_callback(callback: CallbackQuery) -> bool:
    user_id = int(callback.from_user.id)
    username = callback.from_user.username
    DB.upsert_user(user_id=user_id, username=username)

    if user_id not in ADMIN_IDS and DB.is_user_banned(user_id):
        await callback.answer("Доступ запрещен.", show_alert=True)
        return False
    return True


class AddExpenseStates(StatesGroup):
    amount = State()
    category = State()
    comment = State()


class AddFuelStates(StatesGroup):
    liters = State()
    price_per_liter = State()
    mileage = State()


class ReminderStates(StatesGroup):
    reminder_type = State()
    next_date = State()
    amount = State()
    periodicity = State()


class SearchStates(StatesGroup):
    date_start = State()
    date_end = State()
    amount_single = State()
    amount_between_start = State()
    amount_between_end = State()
    comment_text = State()
    combined_query = State()
    results = State()


router = Router()

REMINDER_TASK: asyncio.Task | None = None
BOT_REF: Bot | None = None


@router.message(Command("start"))
async def start_handler(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        print(f"[bot] /start user_id={message.from_user.id}")
        await state.clear()
        await message.answer(
            "Привет! Я бот для учета расходов на автомобиль.",
            reply_markup=get_main_keyboard(),
        )
    except Exception as e:
        print(f"[bot] start_handler error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


@router.message(F.text == MAIN_BUTTON_ADD)
async def add_expense_start(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        print(f"[bot] add_expense_start user_id={message.from_user.id}")
        await state.clear()
        await state.set_state(AddExpenseStates.amount)
        await message.answer("Введите сумму (только цифры):")
    except Exception as e:
        print(f"[bot] add_expense_start error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


@router.message(StateFilter(AddExpenseStates.amount))
async def add_expense_amount(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        text = (message.text or "").strip()
        print(f"[bot] amount user_id={message.from_user.id} text={text!r}")

        if not text.isdigit():
            await message.answer("Сумма должна состоять только из цифр. Введите сумму еще раз:")
            return

        amount = int(text)
        if amount <= 0:
            await message.answer("Сумма должна быть больше 0. Введите сумму еще раз:")
            return

        await state.update_data(amount=amount)
        await state.set_state(AddExpenseStates.category)

        await message.answer(
            "Выберите категорию:",
            reply_markup=get_categories_keyboard(CATEGORIES),
        )
    except Exception as e:
        print(f"[bot] add_expense_amount error: {e}")
        await message.answer("Ошибка при вводе суммы. Попробуй снова.")


@router.callback_query(StateFilter(AddExpenseStates.category), F.data.startswith("cat:"))
async def add_expense_category(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        user_id = callback.from_user.id
        data = callback.data or ""
        category = data.split("cat:", 1)[1]
        category = category.strip()
        print(f"[bot] category user_id={user_id} category={category!r}")

        if category not in CATEGORIES:
            await callback.answer("Неизвестная категория.", show_alert=True)
            return

        await state.update_data(category=category)
        await state.set_state(AddExpenseStates.comment)

        # Не меняем message текстом — просто отправляем новое.
        await callback.message.answer(
            "Введите комментарий или отправьте `Пропустить`.",
            parse_mode="Markdown",
        )
        await callback.answer()
    except Exception as e:
        print(f"[bot] add_expense_category error: {e}")
        await callback.answer("Ошибка. Попробуй снова.", show_alert=True)


@router.message(StateFilter(AddExpenseStates.comment))
async def add_expense_comment(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = message.from_user.id
        text = (message.text or "").strip()
        print(f"[bot] comment user_id={user_id} text={text!r}")

        data = await state.get_data()
        amount = int(data["amount"])
        category = str(data["category"])

        if text.lower() in {"пропустить", "пропущено", "skip"}:
            comment = None
        else:
            comment = text

        # Сохраняем в БД. sqlite3 синхронный, но для MVP достаточно.
        DB.add_expense(
            user_id=user_id,
            amount=amount,
            category=category,
            comment=comment,
        )

        await state.clear()
        currency = DB.get_user_currency(user_id)
        await message.answer(
            f"Расход добавлен!\n{CATEGORY_DISPLAY.get(category, category)}: {_format_amount(amount, currency)}",
            reply_markup=get_main_keyboard(),
        )
    except Exception as e:
        print(f"[bot] add_expense_comment error: {e}")
        await message.answer("Ошибка при сохранении. Попробуй снова.")


@router.message(F.text == MAIN_BUTTON_TODAY)
async def today_stats_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = message.from_user.id
        print(f"[bot] today_stats user_id={user_id}")

        stats = DB.get_today_stats(user_id)
        if stats.total <= 0:
            await message.answer(
                "Статистика за сегодня:\nНет расходов за сегодня.",
                reply_markup=get_main_keyboard(),
            )
            return

        currency = DB.get_user_currency(user_id)
        lines = [f"Статистика за сегодня: {_format_amount(stats.total, currency)}"]
        for cat in sorted(stats.by_category.keys()):
            lines.append(f"{CATEGORY_DISPLAY.get(cat, cat)}: {_format_amount(stats.by_category[cat], currency)}")

        await message.answer("\n".join(lines), reply_markup=get_main_keyboard())
    except Exception as e:
        print(f"[bot] today_stats_handler error: {e}")
        await message.answer("Ошибка при получении статистики. Попробуй позже.")


@router.message(F.text == MAIN_BUTTON_WEEK)
async def week_stats_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = message.from_user.id
        print(f"[bot] week_stats user_id={user_id}")

        stats = DB.get_week_stats(user_id)
        if stats.total <= 0:
            await message.answer(
                "Статистика за неделю:\nНет расходов за последние 7 дней.",
                reply_markup=get_main_keyboard(),
            )
            return

        currency = DB.get_user_currency(user_id)
        lines = [f"Статистика за неделю: {_format_amount(stats.total, currency)}"]
        if not stats.by_day:
            lines.append("Нет данных по дням.")
        else:
            for day, total in stats.by_day:
                lines.append(f"{day}: {_format_amount(total, currency)}")

        await message.answer("\n".join(lines), reply_markup=get_main_keyboard())
    except Exception as e:
        print(f"[bot] week_stats_handler error: {e}")
        await message.answer("Ошибка при получении статистики. Попробуй позже.")


@router.message(F.text == MAIN_BUTTON_PROFILE)
async def profile_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)

        total = DB.get_total_stats(user_id)
        count = DB.get_expenses_count(user_id)
        avg_check = (total / count) if count > 0 else 0
        most_cat = DB.get_most_frequent_category(user_id)

        lines = [
            "👤 Мой профиль",
            f"Всего потрачено: {_format_amount(total, currency)}",
            f"Средний чек: {_format_amount(avg_check, currency)}",
            f"Самая частотная категория: {CATEGORY_DISPLAY.get(most_cat, most_cat) if most_cat else '—'}",
            f"Количество записей в базе: {count}",
        ]
        await message.answer("\n".join(lines), reply_markup=get_main_keyboard())
    except Exception as e:
        print(f"[bot] profile_handler error: {e}")
        await message.answer("Ошибка при получении профиля. Попробуй позже.")


@router.message(Command("currency"))
async def currency_command(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        current = DB.get_user_currency(user_id).upper()
        await message.answer(
            f"Текущая валюта: {current} ({CURRENCY_SYMBOLS.get(current, '')})\nВыберите новую:",
            reply_markup=get_currency_keyboard(CURRENCIES),
        )
    except Exception as e:
        print(f"[bot] currency_command error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


@router.callback_query(F.data.startswith("currency:"))
async def currency_callback(callback: CallbackQuery) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        data = callback.data or ""
        currency = data.split("currency:", 1)[1].strip().upper()
        user_id = int(callback.from_user.id)
        if currency not in CURRENCIES:
            await callback.answer("Неизвестная валюта.", show_alert=True)
            return
        DB.set_user_currency(user_id, currency)
        await callback.message.answer(
            f"Валюта обновлена: {currency} ({CURRENCY_SYMBOLS.get(currency, '')})",
            reply_markup=get_main_keyboard(),
        )
        await callback.answer()
    except Exception as e:
        print(f"[bot] currency_callback error: {e}")
        await callback.answer("Ошибка.", show_alert=True)


@router.message(F.text == MAIN_BUTTON_ADD_FUEL)
async def add_fuel_start(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        await state.clear()
        await state.set_state(AddFuelStates.liters)
        await message.answer("⛽ Введите количество литров:")
    except Exception as e:
        print(f"[bot] add_fuel_start error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


@router.message(StateFilter(AddFuelStates.liters))
async def add_fuel_liters(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_float_ru(message.text or "")
        if value is None or value <= 0:
            await message.answer("Литры должны быть числом больше 0. Введите еще раз:")
            return
        await state.update_data(liters=value)
        await state.set_state(AddFuelStates.price_per_liter)
        await message.answer("Введите цену за 1 литр:")
    except Exception as e:
        print(f"[bot] add_fuel_liters error: {e}")
        await message.answer("Ошибка при вводе литров. Попробуй снова.")


@router.message(StateFilter(AddFuelStates.price_per_liter))
async def add_fuel_price(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_float_ru(message.text or "")
        if value is None or value <= 0:
            await message.answer("Цена должна быть числом больше 0. Введите еще раз:")
            return
        await state.update_data(price_per_liter=value)
        await state.set_state(AddFuelStates.mileage)
        await message.answer("Введите текущий пробег (км):")
    except Exception as e:
        print(f"[bot] add_fuel_price error: {e}")
        await message.answer("Ошибка при вводе цены. Попробуй снова.")


@router.message(StateFilter(AddFuelStates.mileage))
async def add_fuel_mileage(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_float_ru(message.text or "")
        if value is None or value <= 0:
            await message.answer("Пробег должен быть числом больше 0. Введите еще раз:")
            return

        data = await state.get_data()
        user_id = int(message.from_user.id)
        liters = float(data["liters"])
        price_per_liter = float(data["price_per_liter"])
        mileage = float(value)
        currency = DB.get_user_currency(user_id)

        try:
            FUEL_DB.add_fuel_log(
                user_id=user_id,
                liters=liters,
                price_per_liter=price_per_liter,
                mileage=mileage,
                car_model=FUEL_CAR_MODEL_DEFAULT,
            )
        except ValueError as ve:
            await message.answer(f"Ошибка: {ve}\nВведите пробег еще раз (должен быть больше предыдущего).")
            return

        total_cost_rub = int(round(liters * price_per_liter))
        await state.clear()
        await message.answer(
            f"Заправка добавлена!\nСумма: {_format_amount(total_cost_rub, currency)}",
            reply_markup=get_main_keyboard(),
        )
    except Exception as e:
        print(f"[bot] add_fuel_mileage error: {e}")
        await message.answer("Ошибка при добавлении заправки. Попробуй снова.")


@router.message(F.text == MAIN_BUTTON_FUEL_STATS)
async def fuel_stats_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        stats = FUEL_DB.get_fuel_stats(user_id)

        total_cost = int(stats.get("total_cost") or 0)
        avg_last_5 = stats.get("avg_last_5")
        min_cons = stats.get("min_consumption")
        max_cons = stats.get("max_consumption")

        lines: list[str] = ["📊 Расход топлива"]
        lines.append(f"Итоговая сумма на топливо: {_format_amount(total_cost, currency)}")
        if avg_last_5 is None:
            lines.append("Недостаточно данных для расчёта расхода на 100 км.")
        else:
            lines.append(f"Средний расход за последние 5 заправок: {avg_last_5:.2f} л/100 км")
            lines.append(f"Самый экономичный расход: {min_cons:.2f} л/100 км")
            lines.append(f"Самый высокий расход: {max_cons:.2f} л/100 км")

        await message.answer("\n".join(lines), reply_markup=get_main_keyboard())
    except Exception as e:
        print(f"[bot] fuel_stats_handler error: {e}")
        await message.answer("Ошибка при получении статистики топлива. Попробуй позже.")


@router.message(F.text == MAIN_BUTTON_GRAPH_EXPENSES)
async def expenses_graph_request(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        await message.answer("Выберите период для графика:", reply_markup=get_period_keyboard())
    except Exception as e:
        print(f"[bot] expenses_graph_request error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


@router.callback_query(F.data.startswith("period:"))
async def expenses_graph_period(callback: CallbackQuery) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        period = (callback.data or "").split("period:", 1)[1].strip()
        user_id = int(callback.from_user.id)
        currency = DB.get_user_currency(user_id)

        if period == "week":
            start_date, end_date = _this_week_range()
        elif period == "month":
            start_date, end_date = _this_month_range()
        else:
            start_date, end_date = None, None

        data = DB.get_expenses_by_category_between(user_id, start_date, end_date)
        if not data:
            await callback.message.answer("Нет данных для выбранного периода.", reply_markup=get_main_keyboard())
            await callback.answer()
            return

        import matplotlib.pyplot as plt
        import pandas as pd

        df = pd.DataFrame({"category": list(data.keys()), "total": list(data.values())})
        df = df.sort_values("total", ascending=False)
        df["total"] = df["total"].apply(lambda v: _convert_amount_rub(v, currency))

        plt.figure(figsize=(10, 6))
        plt.bar(df["category"].apply(lambda c: CATEGORY_DISPLAY.get(c, c)), df["total"])
        plt.title("Расходы по категориям")
        plt.xlabel("Категория")
        plt.ylabel(f"Сумма ({currency})")
        plt.xticks(rotation=20, ha="right")
        plt.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=160)
        plt.close()
        buf.seek(0)

        photo = BufferedInputFile(buf.getvalue(), filename="expenses_graph.png")
        await callback.message.answer_photo(photo=photo)
        await callback.answer()
    except Exception as e:
        print(f"[bot] expenses_graph_period error: {e}")
        await callback.answer("Ошибка при построении графика.", show_alert=True)


@router.message(F.text == MAIN_BUTTON_DYNAMIC_EXPENSES)
async def expenses_dynamic_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        start_date = (date.today() - timedelta(days=29)).isoformat()
        end_date = date.today().isoformat()

        points = DB.get_expenses_by_day_between(user_id, start_date, end_date)
        if not points:
            await message.answer("Нет данных для динамики за последние 30 дней.", reply_markup=get_main_keyboard())
            return

        import matplotlib.pyplot as plt
        import pandas as pd

        df = pd.DataFrame(points, columns=["date", "total"])
        df["total"] = df["total"].apply(lambda v: _convert_amount_rub(v, currency))
        plt.figure(figsize=(10, 6))
        plt.plot(df["date"], df["total"], marker="o")
        plt.title("Динамика расходов по дням (30 дней)")
        plt.xlabel("Дата")
        plt.ylabel(f"Сумма ({currency})")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=160)
        plt.close()
        buf.seek(0)

        photo = BufferedInputFile(buf.getvalue(), filename="expenses_dynamic.png")
        await message.answer_photo(photo=photo)
    except Exception as e:
        print(f"[bot] expenses_dynamic_handler error: {e}")
        await message.answer("Ошибка при построении графика. Попробуй позже.")


@router.message(F.text == MAIN_BUTTON_GRAPH_FUEL)
async def fuel_consumption_graph_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        points = FUEL_DB.get_fuel_points_for_graph(user_id)
        if len(points) < 2:
            await message.answer("Недостаточно данных для графика расхода топлива.", reply_markup=get_main_keyboard())
            return

        xs = [p[0] for p in points]
        ys = [p[1] for p in points]

        # Линейная регрессия y = a*x + b (без numpy).
        x_mean = sum(xs) / len(xs)
        y_mean = sum(ys) / len(ys)
        denom = sum((x - x_mean) ** 2 for x in xs) or 1.0
        a = sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys)) / denom
        b = y_mean - a * x_mean

        trend_y = [a * x + b for x in xs]

        import matplotlib.pyplot as plt

        plt.figure(figsize=(10, 6))
        plt.scatter(xs, ys, label="Заправки")
        plt.plot(xs, trend_y, color="red", linestyle="--", label="Тренд")
        plt.title("Пробег vs расход на 100 км")
        plt.xlabel("Пробег (км)")
        plt.ylabel("Расход (л/100 км)")
        plt.legend()
        plt.tight_layout()

        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=160)
        plt.close()
        buf.seek(0)

        photo = BufferedInputFile(buf.getvalue(), filename="fuel_consumption_graph.png")
        await message.answer_photo(photo=photo)
    except Exception as e:
        print(f"[bot] fuel_consumption_graph_handler error: {e}")
        await message.answer("Ошибка при построении графика. Попробуй позже.")


def _excel_to_bytes(workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


@router.message(F.text == MAIN_BUTTON_EXPORT_EXCEL)
async def export_excel_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)

        import openpyxl

        wb = openpyxl.Workbook()
        ws_expenses = wb.active
        ws_expenses.title = "Расходы"
        ws_expenses.append(["id", "user_id", "amount_rub", "category", "comment", "date"])

        for r in DB.get_expenses_for_export(user_id):
            ws_expenses.append([r["id"], r["user_id"], r["amount"], r["category"], r["comment"], r["date"]])

        ws_fuel = wb.create_sheet("Топливо")
        ws_fuel.append(["id", "user_id", "liters", "price_per_liter", "total_cost_rub", "mileage", "car_model", "date"])
        for r in FUEL_DB.get_fuel_logs_for_export(user_id):
            ws_fuel.append([r["id"], r["user_id"], r["liters"], r["price_per_liter"], r["total_cost"], r["mileage"], r["car_model"], r["date"]])

        data = _excel_to_bytes(wb)
        bio = BytesIO(data)
        bio.name = "expenses_export.xlsx"

        await message.answer_document(document=BufferedInputFile(data, filename="expenses_export.xlsx"))
    except Exception as e:
        print(f"[bot] export_excel_handler error: {e}")
        await message.answer("Ошибка при экспорте Excel. Попробуй позже.")


@router.message(Command("connect_google"))
async def connect_google_handler(message: Message) -> None:
    """
    Подключение и обновление Google Sheets (PROMPT #6).
    Ожидается: credentials.json в корне проекта.
    """
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        import os

        credentials_path = os.path.join(os.getcwd(), "credentials.json")
        if not os.path.exists(credentials_path):
            # Railway/Fly.io часто удобнее хранить JSON ключа в env, чем как файл.
            # Если задана переменная `GOOGLE_CREDENTIALS_JSON`, создаём credentials.json на диске.
            credentials_json = os.getenv("GOOGLE_CREDENTIALS_JSON", "").strip()
            if credentials_json:
                try:
                    from pathlib import Path

                    Path(credentials_path).write_text(credentials_json, encoding="utf-8")
                except Exception as e:
                    await message.answer(
                        "Не удалось создать `credentials.json` из `GOOGLE_CREDENTIALS_JSON`.\n"
                        f"Ошибка: {e}",
                        reply_markup=get_main_keyboard(),
                    )
                    return
            else:
                await message.answer(
                    "Не найден файл `credentials.json`.\n"
                    "Варианты:\n"
                    "1) Положи `credentials.json` в корень проекта.\n"
                    "2) Или задай `GOOGLE_CREDENTIALS_JSON` (переменная окружения) на хостинге и повтори /connect_google.",
                    reply_markup=get_main_keyboard(),
                )
                return

        import gspread
        from oauth2client.service_account import ServiceAccountCredentials

        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(credentials_path, scope)
        gc = gspread.authorize(creds)

        spreadsheet_id = DB.get_user_google_spreadsheet_id(user_id)
        if spreadsheet_id:
            sh = gc.open_by_key(spreadsheet_id)
        else:
            sh = gc.create(f"car_cost_bot_{user_id}")
            DB.set_user_google_spreadsheet_id(user_id, sh.id)

        # Sheet 1: Расходы
        try:
            ws_expenses = sh.worksheet("Расходы")
        except Exception:
            ws_expenses = sh.add_worksheet(title="Расходы", rows=2000, cols=8)

        ws_expenses.clear()
        ws_expenses.append_row(["id", "user_id", "amount_rub", "category", "comment", "date"])
        for r in DB.get_expenses_for_export(user_id):
            ws_expenses.append_row([r["id"], r["user_id"], r["amount"], r["category"], r["comment"], r["date"]])

        # Sheet 2: Топливо
        try:
            ws_fuel = sh.worksheet("Топливо")
        except Exception:
            ws_fuel = sh.add_worksheet(title="Топливо", rows=2000, cols=8)

        ws_fuel.clear()
        ws_fuel.append_row(["id", "user_id", "liters", "price_per_liter", "total_cost_rub", "mileage", "car_model", "date"])
        for r in FUEL_DB.get_fuel_logs_for_export(user_id):
            ws_fuel.append_row([r["id"], r["user_id"], r["liters"], r["price_per_liter"], r["total_cost"], r["mileage"], r["car_model"], r["date"]])

        await message.answer(f"Google Sheets обновлены. ID таблицы: {sh.id}", reply_markup=get_main_keyboard())
    except Exception as e:
        print(f"[bot] connect_google_handler error: {e}")
        await message.answer("Ошибка подключения к Google Sheets. Проверь credentials.json и попробуй снова.")


@router.message(F.text == MAIN_BUTTON_REMINDERS)
async def reminders_add_start(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        await state.clear()
        await state.set_state(ReminderStates.reminder_type)
        await message.answer("Выберите тип напоминания:", reply_markup=get_reminder_types_keyboard())
    except Exception as e:
        print(f"[bot] reminders_add_start error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


@router.callback_query(F.data.startswith("rem_type:"))
async def reminders_pick_type(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        data = callback.data or ""
        reminder_type = data.split("rem_type:", 1)[1].strip()
        user_id = int(callback.from_user.id)

        await state.update_data(reminder_type=reminder_type)
        await state.set_state(ReminderStates.next_date)

        await callback.message.answer(
            "Введите дату следующего события в формате YYYY-MM-DD (например, 2026-05-10):"
        )
        await callback.answer()
    except Exception as e:
        print(f"[bot] reminders_pick_type error: {e}")
        await callback.answer("Ошибка.", show_alert=True)


@router.message(StateFilter(ReminderStates.next_date))
async def reminders_set_date(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        text = (message.text or "").strip()
        try:
            # валидируем ISO
            date.fromisoformat(text)
        except ValueError:
            await message.answer("Дата должна быть в формате YYYY-MM-DD. Введите еще раз:")
            return

        await state.update_data(next_date=text)
        await state.set_state(ReminderStates.amount)
        await message.answer("Введите сумму (примерно, целым числом):")
    except Exception as e:
        print(f"[bot] reminders_set_date error: {e}")
        await message.answer("Ошибка при вводе даты. Попробуй снова.")


@router.message(StateFilter(ReminderStates.amount))
async def reminders_set_amount(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_int_ru(message.text or "")
        if value is None or value <= 0:
            await message.answer("Сумма должна быть целым числом > 0. Введите еще раз:")
            return

        await state.update_data(amount=value)
        await state.set_state(ReminderStates.periodicity)
        await message.answer("Выберите периодичность:", reply_markup=get_reminder_periodicity_keyboard())
    except Exception as e:
        print(f"[bot] reminders_set_amount error: {e}")
        await message.answer("Ошибка при вводе суммы. Попробуй снова.")


@router.callback_query(F.data.startswith("rem_period:"))
async def reminders_pick_periodicity(callback: CallbackQuery, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        periodicity = (callback.data or "").split("rem_period:", 1)[1].strip()
        user_id = int(callback.from_user.id)

        data = await state.get_data()
        reminder_type = str(data["reminder_type"])
        next_date = str(data["next_date"])
        amount = int(data["amount"])

        DB.add_reminder(
            user_id=user_id,
            reminder_type=reminder_type,
            next_date=next_date,
            amount=amount,
            periodicity=periodicity,
        )
        await state.clear()

        currency = DB.get_user_currency(user_id)
        await callback.message.answer(
            f"Напоминание добавлено!\n{reminder_type}: {next_date}\nСумма: {_format_amount(amount, currency)}\nПериодичность: {REMINDER_PERIODICITIES.get(periodicity, periodicity)}",
            reply_markup=get_main_keyboard(),
        )
        await callback.answer()
    except Exception as e:
        print(f"[bot] reminders_pick_periodicity error: {e}")
        await callback.answer("Ошибка.", show_alert=True)


@router.message(F.text == MAIN_BUTTON_REMINDERS_LIST)
async def reminders_list_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        rows = DB.get_active_reminders(user_id)
        if not rows:
            await message.answer("Активных напоминаний нет.", reply_markup=get_main_keyboard())
            return

        lines = ["📋 Список напоминаний:"]
        for r in rows:
            lines.append(
                f"#{r['id']} | {r['reminder_type']} | {r['next_date']} | {_format_amount(r['amount'], currency)} | {REMINDER_PERIODICITIES.get(r['periodicity'], r['periodicity'])}"
            )
        await message.answer("\n".join(lines), reply_markup=get_main_keyboard())
    except Exception as e:
        print(f"[bot] reminders_list_handler error: {e}")
        await message.answer("Ошибка при получении списка напоминаний. Попробуй позже.")


@router.message(Command("users"))
async def admin_users_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return

        rows = DB.conn.execute(
            """
            SELECT u.user_id, u.username, u.is_banned, u.currency,
                   COALESCE(SUM(e.amount), 0) AS total_spent,
                   COALESCE(COUNT(e.id), 0) AS ops_count
            FROM users u
            LEFT JOIN expenses e ON e.user_id = u.user_id
            GROUP BY u.user_id
            ORDER BY u.user_id ASC
            """
        ).fetchall()

        if not rows:
            await message.answer("Пользователей пока нет.")
            return

        lines = ["/users — список пользователей"]
        for r in rows:
            uid = int(r["user_id"])
            username = r["username"] or ""
            currency = r["currency"] or "RUB"
            total = int(r["total_spent"] or 0)
            ops = int(r["ops_count"] or 0)
            lines.append(
                f"{uid} | @{username} | banned={r['is_banned']} | ops={ops} | total={_format_amount(total, currency)}"
            )
        await message.answer("\n".join(lines))
    except Exception as e:
        print(f"[bot] admin_users_handler error: {e}")
        await message.answer("Ошибка в /users.")


@router.message(F.text.startswith("/broadcast"))
async def admin_broadcast_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return

        text = message.text or ""
        # /broadcast [текст]
        parts = text.split(" ", 1)
        if len(parts) < 2:
            await message.answer("Использование: /broadcast [текст]")
            return
        payload = parts[1].strip()

        user_ids = DB.conn.execute(
            "SELECT user_id FROM users WHERE is_banned = 0"
        ).fetchall()
        ok_count = 0
        for r in user_ids:
            uid = int(r["user_id"])
            await message.bot.send_message(uid, payload)
            ok_count += 1
        await message.answer(f"Рассылка завершена. Сообщений отправлено: {ok_count}")
    except Exception as e:
        print(f"[bot] admin_broadcast_handler error: {e}")
        await message.answer("Ошибка в /broadcast.")


@router.message(Command("stats_all"))
async def admin_stats_all_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return

        total_row = DB.conn.execute("SELECT COALESCE(SUM(amount), 0) AS total FROM expenses").fetchone()
        total = int(total_row["total"] or 0)
        currency = "RUB"

        rows = DB.conn.execute(
            """
            SELECT category, COALESCE(SUM(amount), 0) AS total
            FROM expenses
            GROUP BY category
            ORDER BY total DESC
            LIMIT 10
            """
        ).fetchall()

        lines = ["📊 Общая статистика по всем пользователям", f"Итого: {_format_amount(total, currency)}"]
        for r in rows:
            cat = str(r["category"])
            lines.append(f"{CATEGORY_DISPLAY.get(cat, cat)}: {_format_amount(int(r['total'] or 0), currency)}")

        await message.answer("\n".join(lines))
    except Exception as e:
        print(f"[bot] admin_stats_all_handler error: {e}")
        await message.answer("Ошибка в /stats_all.")


@router.message(Command("top_users"))
async def admin_top_users_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return

        rows = DB.conn.execute(
            """
            SELECT u.user_id, u.username, COALESCE(COUNT(e.id), 0) AS ops
            FROM users u
            LEFT JOIN expenses e ON e.user_id = u.user_id
            GROUP BY u.user_id
            ORDER BY ops DESC
            LIMIT 10
            """
        ).fetchall()

        lines = ["🏆 Топ 10 пользователей по количеству операций"]
        for r in rows:
            uid = int(r["user_id"])
            username = r["username"] or ""
            ops = int(r["ops"] or 0)
            lines.append(f"{uid} | @{username} | ops={ops}")

        await message.answer("\n".join(lines))
    except Exception as e:
        print(f"[bot] admin_top_users_handler error: {e}")
        await message.answer("Ошибка в /top_users.")


@router.message(Command("export_all"))
async def admin_export_all_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return

        import openpyxl

        wb = openpyxl.Workbook()
        ws_expenses = wb.active
        ws_expenses.title = "Расходы"
        ws_expenses.append(["id", "user_id", "amount_rub", "category", "comment", "date"])

        for r in DB.get_expenses_for_export_all():
            ws_expenses.append([r["id"], r["user_id"], r["amount"], r["category"], r["comment"], r["date"]])

        ws_fuel = wb.create_sheet("Топливо")
        ws_fuel.append(["id", "user_id", "liters", "price_per_liter", "total_cost_rub", "mileage", "car_model", "date"])
        for r in FUEL_DB.get_fuel_logs_for_export_all():
            ws_fuel.append([r["id"], r["user_id"], r["liters"], r["price_per_liter"], r["total_cost"], r["mileage"], r["car_model"], r["date"]])

        data = _excel_to_bytes(wb)
        await message.answer_document(document=BufferedInputFile(data, filename="expenses_export_all.xlsx"))
    except Exception as e:
        print(f"[bot] admin_export_all_handler error: {e}")
        await message.answer("Ошибка в /export_all.")


def _parse_user_id_arg(message: Message) -> int | None:
    text = message.text or ""
    parts = text.split()
    if len(parts) < 2:
        return None
    try:
        return int(parts[1])
    except ValueError:
        return None


@router.message(Command("ban"))
async def admin_ban_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return
        target_id = _parse_user_id_arg(message)
        if target_id is None:
            await message.answer("Использование: /ban [user_id]")
            return
        DB.set_ban(target_id, True)
        await message.answer(f"Пользователь заблокирован: {target_id}")
    except Exception as e:
        print(f"[bot] admin_ban_handler error: {e}")
        await message.answer("Ошибка в /ban.")


@router.message(Command("unban"))
async def admin_unban_handler(message: Message) -> None:
    try:
        user_id = int(message.from_user.id)
        if user_id not in ADMIN_IDS:
            await message.answer("Только для админа.")
            return
        target_id = _parse_user_id_arg(message)
        if target_id is None:
            await message.answer("Использование: /unban [user_id]")
            return
        DB.set_ban(target_id, False)
        await message.answer(f"Пользователь разблокирован: {target_id}")
    except Exception as e:
        print(f"[bot] admin_unban_handler error: {e}")
        await message.answer("Ошибка в /unban.")


@router.message(Command("search"))
async def search_command_handler(message: Message) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        await message.answer("Выберите тип поиска:", reply_markup=get_search_menu_keyboard())
    except Exception as e:
        print(f"[bot] search_command_handler error: {e}")
        await message.answer("Ошибка. Попробуй позже.")


async def _send_search_page(chat_id: int, user_id: int, currency: str, state: FSMContext, search_params: dict, page: int, bot: Bot, edit_message_text: str | None = None) -> None:
    limit = SEARCH_PER_PAGE
    offset = (page - 1) * limit
    res = DB.search_expenses(
        user_id=user_id,
        start_date=search_params.get("start_date"),
        end_date=search_params.get("end_date"),
        category=search_params.get("category"),
        amount_min=search_params.get("amount_min"),
        amount_max=search_params.get("amount_max"),
        comment_contains=search_params.get("comment_contains"),
        limit=limit,
        offset=offset,
    )
    if res.total == 0:
        text = "По запросу ничего не найдено."
        if edit_message_text is not None and state:
            await bot.send_message(chat_id, text)
        else:
            await bot.send_message(chat_id, text)
        await state.set_state(SearchStates.results)
        await state.update_data(search_params=search_params, page=1)
        return

    total_pages = max(1, (res.total + limit - 1) // limit)
    page = min(max(1, page), total_pages)

    lines: list[str] = []
    for item in res.items:
        d, cat, amount_rub, comment = item
        cat_disp = CATEGORY_DISPLAY.get(cat, cat)
        comment_text = comment if comment else "-"
        lines.append(f"📅 {d} | {cat_disp} | {_format_amount(amount_rub, currency)} | {comment_text}")

    header = f"Результаты (страница {page}/{total_pages}):"
    kb = get_pagination_keyboard(page=page, total_pages=total_pages)

    if edit_message_text is not None:
        # не используем редактирование в этом MVP
        await bot.send_message(chat_id, header + "\n" + "\n".join(lines), reply_markup=kb)
    else:
        await bot.send_message(chat_id, header + "\n" + "\n".join(lines), reply_markup=kb)

    await state.set_state(SearchStates.results)
    await state.update_data(search_params=search_params, page=page)


@router.callback_query(F.data.startswith("search:"))
async def search_menu_callback(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        user_id = int(callback.from_user.id)
        currency = DB.get_user_currency(user_id)
        chat_id = callback.message.chat.id

        key = (callback.data or "").split("search:", 1)[1]
        await callback.answer()

        if key == "date_today":
            start = _today_iso()
            end = _today_iso()
            params = {"start_date": start, "end_date": end, "category": None, "amount_min": None, "amount_max": None, "comment_contains": None}
            await _send_search_page(chat_id, user_id, currency, state, params, page=1, bot=bot)
            return
        if key == "date_yesterday":
            start = _yesterday_iso()
            end = _yesterday_iso()
            params = {"start_date": start, "end_date": end, "category": None, "amount_min": None, "amount_max": None, "comment_contains": None}
            await _send_search_page(chat_id, user_id, currency, state, params, page=1, bot=bot)
            return
        if key == "date_week":
            start, end = _this_week_range()
            params = {"start_date": start, "end_date": end, "category": None, "amount_min": None, "amount_max": None, "comment_contains": None}
            await _send_search_page(chat_id, user_id, currency, state, params, page=1, bot=bot)
            return
        if key == "date_month":
            start, end = _this_month_range()
            params = {"start_date": start, "end_date": end, "category": None, "amount_min": None, "amount_max": None, "comment_contains": None}
            await _send_search_page(chat_id, user_id, currency, state, params, page=1, bot=bot)
            return

        if key == "date_range":
            await state.clear()
            await state.set_state(SearchStates.date_start)
            await bot.send_message(chat_id, "Введите ДАТУ НАЧАЛА (YYYY-MM-DD):")
            return

        if key == "category":
            # Категория из списка
            categories = CATEGORIES
            buttons = []
            row = []
            for idx, cat in enumerate(categories, start=1):
                row.append(InlineKeyboardButton(text=cat, callback_data=f"search_cat:{cat}"))
                if idx % 2 == 0:
                    buttons.append(row)
                    row = []
            if row:
                buttons.append(row)
            markup = InlineKeyboardMarkup(inline_keyboard=buttons)

            await state.clear()
            await state.set_state(SearchStates.results)
            await bot.send_message(chat_id, "Выберите категорию:", reply_markup=markup)
            return

        if key == "amount_gt":
            await state.clear()
            await state.set_state(SearchStates.amount_single)
            await bot.send_message(chat_id, "Введите число X для поиска: больше чем X (только целое):")
            return

        if key == "amount_lt":
            await state.clear()
            await state.set_state(SearchStates.amount_single)
            await bot.send_message(chat_id, "Введите число X для поиска: меньше чем X (только целое):")
            # отметим тип через state data
            await state.update_data(search_mode="amount_lt")
            return

        if key == "amount_between":
            await state.clear()
            await state.set_state(SearchStates.amount_between_start)
            await bot.send_message(chat_id, "Введите X (начало диапазона):")
            return

        if key == "comment":
            await state.clear()
            await state.set_state(SearchStates.comment_text)
            await bot.send_message(chat_id, "Введите текст для поиска по комментарию:")
            return

        if key == "combined":
            await state.clear()
            await state.set_state(SearchStates.combined_query)
            await bot.send_message(chat_id, "Введите запрос (например: 'топливо за прошлый месяц от 2000 до 5000'):")
            return

        await bot.send_message(chat_id, "Опция пока не поддержана.")
    except Exception as e:
        print(f"[bot] search_menu_callback error: {e}")
        await callback.answer("Ошибка.", show_alert=True)


@router.callback_query(F.data.startswith("search_cat:"))
async def search_category_callback(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        user_id = int(callback.from_user.id)
        currency = DB.get_user_currency(user_id)
        chat_id = callback.message.chat.id
        cat = (callback.data or "").split("search_cat:", 1)[1].strip()
        params = {
            "start_date": None,
            "end_date": None,
            "category": cat,
            "amount_min": None,
            "amount_max": None,
            "comment_contains": None,
        }
        await _send_search_page(chat_id, user_id, currency, state, params, page=1, bot=bot)
        await callback.answer()
    except Exception as e:
        print(f"[bot] search_category_callback error: {e}")
        await callback.answer("Ошибка.", show_alert=True)


@router.callback_query(F.data.startswith("search_page:"))
async def search_pagination_callback(callback: CallbackQuery, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_callback(callback):
            return
        user_id = int(callback.from_user.id)
        currency = DB.get_user_currency(user_id)
        chat_id = callback.message.chat.id

        page = int((callback.data or "").split("search_page:", 1)[1].strip())
        data = await state.get_data()
        search_params = data.get("search_params") or {}
        await _send_search_page(chat_id, user_id, currency, state, search_params, page=page, bot=bot)
        await callback.answer()
    except Exception as e:
        print(f"[bot] search_pagination_callback error: {e}")
        await callback.answer("Ошибка.", show_alert=True)


@router.message(StateFilter(SearchStates.date_start))
async def search_date_start_input(message: Message, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        text = (message.text or "").strip()
        date_from = date.fromisoformat(text)
        await state.update_data(start_date=date_from.isoformat())
        await state.set_state(SearchStates.date_end)
        await message.answer("Введите ДАТУ КОНЦА (YYYY-MM-DD):")
    except Exception:
        await message.answer("Неверный формат даты. Введите YYYY-MM-DD.")


@router.message(StateFilter(SearchStates.date_end))
async def search_date_end_input(message: Message, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        text = (message.text or "").strip()
        date_to = date.fromisoformat(text)
        data = await state.get_data()
        start_date = data.get("start_date")
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        params = {
            "start_date": start_date,
            "end_date": date_to.isoformat(),
            "category": None,
            "amount_min": None,
            "amount_max": None,
            "comment_contains": None,
        }
        await _send_search_page(message.chat.id, user_id, currency, state, params, page=1, bot=bot)
    except Exception as e:
        print(f"[bot] search_date_end_input error: {e}")
        await message.answer("Ошибка. Проверь дату и попробуй снова.")


@router.message(StateFilter(SearchStates.amount_single))
async def search_amount_single_input(message: Message, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_int_ru(message.text or "")
        if value is None:
            await message.answer("Введите целое число X.")
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        data = await state.get_data()
        mode = data.get("search_mode", "amount_gt")

        if mode == "amount_lt":
            params = {
                "start_date": None,
                "end_date": None,
                "category": None,
                "amount_min": None,
                "amount_max": value,
                "comment_contains": None,
            }
        else:
            params = {
                "start_date": None,
                "end_date": None,
                "category": None,
                "amount_min": value,
                "amount_max": None,
                "comment_contains": None,
            }
        await _send_search_page(message.chat.id, user_id, currency, state, params, page=1, bot=bot)
    except Exception as e:
        print(f"[bot] search_amount_single_input error: {e}")
        await message.answer("Ошибка. Попробуй снова.")


@router.message(StateFilter(SearchStates.amount_between_start))
async def search_amount_between_start_input(message: Message, state: FSMContext) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_int_ru(message.text or "")
        if value is None:
            await message.answer("Введите целое число X.")
            return
        await state.update_data(amount_between_start=value)
        await state.set_state(SearchStates.amount_between_end)
        await message.answer("Введите Y (конец диапазона):")
    except Exception as e:
        print(f"[bot] search_amount_between_start_input error: {e}")
        await message.answer("Ошибка. Попробуй снова.")


@router.message(StateFilter(SearchStates.amount_between_end))
async def search_amount_between_end_input(message: Message, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        value = _parse_int_ru(message.text or "")
        if value is None:
            await message.answer("Введите целое число Y.")
            return
        data = await state.get_data()
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)

        x = int(data["amount_between_start"])
        y = value
        params = {
            "start_date": None,
            "end_date": None,
            "category": None,
            "amount_min": min(x, y),
            "amount_max": max(x, y),
            "comment_contains": None,
        }
        await _send_search_page(message.chat.id, user_id, currency, state, params, page=1, bot=bot)
    except Exception as e:
        print(f"[bot] search_amount_between_end_input error: {e}")
        await message.answer("Ошибка. Попробуй снова.")


@router.message(StateFilter(SearchStates.comment_text))
async def search_comment_input(message: Message, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        text = (message.text or "").strip()
        if not text:
            await message.answer("Введите текст.")
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        params = {
            "start_date": None,
            "end_date": None,
            "category": None,
            "amount_min": None,
            "amount_max": None,
            "comment_contains": text,
        }
        await _send_search_page(message.chat.id, user_id, currency, state, params, page=1, bot=bot)
    except Exception as e:
        print(f"[bot] search_comment_input error: {e}")
        await message.answer("Ошибка. Попробуй снова.")


def _parse_combined_search_query(query: str) -> dict[str, str | int | None]:
    """
    Эвристический парсер для комбинированного поиска (PROMPT #9).
    """
    q = (query or "").lower()
    start_date: str | None = None
    end_date: str | None = None
    category: str | None = None
    amount_min: int | None = None
    amount_max: int | None = None
    comment_contains: str | None = None

    for c in CATEGORIES:
        if c.lower() in q:
            category = c
            break

    # Даты (ключевые слова)
    if "сегодня" in q:
        start_date = _today_iso()
        end_date = _today_iso()
    elif "вчера" in q:
        start_date = _yesterday_iso()
        end_date = _yesterday_iso()
    elif "эту неделю" in q or "на этой неделе" in q:
        start_date, end_date = _this_week_range()
    elif "этот месяц" in q:
        start_date, end_date = _this_month_range()
    elif "прошлый месяц" in q:
        end = date.today() - timedelta(days=30)
        start = date.today() - timedelta(days=60)
        start_date, end_date = start.isoformat(), end.isoformat()

    # Две даты формата YYYY-MM-DD
    dates = re.findall(r"\d{4}-\d{2}-\d{2}", q)
    if len(dates) >= 2:
        start_date = dates[0]
        end_date = dates[1]

    # Суммы
    m1 = re.search(r"больше\s*чем\s*(\d+)", q)
    if m1:
        amount_min = int(m1.group(1))
    m2 = re.search(r"меньше\s*чем\s*(\d+)", q)
    if m2:
        amount_max = int(m2.group(1))
    m3 = re.search(r"от\s*(\d+)", q)
    if m3:
        amount_min = int(m3.group(1))
    m4 = re.search(r"до\s*(\d+)", q)
    if m4:
        amount_max = int(m4.group(1))

    m5 = re.search(r"(\d+)\s*-\s*(\d+)", q)
    if m5:
        a = int(m5.group(1))
        b = int(m5.group(2))
        amount_min, amount_max = min(a, b), max(a, b)

    # комментарий
    if "коммент" in q or "в комментарии" in q:
        # простая эвристика: берём часть после двоеточия/слова "коммент"
        if ":" in query:
            comment_contains = query.split(":", 1)[1].strip()
        else:
            comment_contains = query

    return {
        "start_date": start_date,
        "end_date": end_date,
        "category": category,
        "amount_min": amount_min,
        "amount_max": amount_max,
        "comment_contains": comment_contains,
    }


@router.message(StateFilter(SearchStates.combined_query))
async def search_combined_input(message: Message, state: FSMContext, bot: Bot) -> None:
    try:
        if not await ensure_user_not_banned_message(message):
            return
        user_id = int(message.from_user.id)
        currency = DB.get_user_currency(user_id)
        query = (message.text or "").strip()
        if not query:
            await message.answer("Введите запрос.")
            return
        params = _parse_combined_search_query(query)
        await _send_search_page(message.chat.id, user_id, currency, state, params, page=1, bot=bot)
    except Exception as e:
        print(f"[bot] search_combined_input error: {e}")
        await message.answer("Ошибка при поиске. Попробуй снова.")


async def reminder_loop() -> None:
    """
    PROMPT #7: каждый день проверяет напоминания в окно [сегодня..завтра] (в 10:00).
    """
    global REMINDER_TASK
    last_run_for: str | None = None
    while True:
        try:
            now = datetime.now()
            run_time = datetime.combine(now.date(), datetime.min.time()).replace(hour=10, minute=0, second=0, microsecond=0)
            if now >= run_time:
                run_time = run_time + timedelta(days=1)
            await asyncio.sleep(max(0, (run_time - now).total_seconds()))

            today = date.today()
            tomorrow = today + timedelta(days=1)
            today_str = today.isoformat()
            tomorrow_str = tomorrow.isoformat()

            # предотвращаем двойной запуск при лаге таймера
            if last_run_for == today_str:
                continue
            last_run_for = today_str

            due = DB.get_due_reminders(today_str, tomorrow_str)
            if not due:
                continue
            bot = BOT_REF
            if bot is None:
                print("[bot] reminder_loop: BOT_REF is None, пропускаем отправку")
                continue

            # Отправляем уведомления и сдвигаем/деактивируем напоминание только если событие сегодня.
            for r in due:
                try:
                    uid = int(r["user_id"])
                    if uid not in ADMIN_IDS and DB.is_user_banned(uid):
                        continue
                    currency = DB.get_user_currency(uid)

                    next_date = str(r["next_date"])
                    prefix = "Сегодня" if next_date == today_str else "Завтра"
                    rem_type = str(r["reminder_type"])
                    amount_rub = int(r["amount"])
                    periodicity = str(r["periodicity"])

                    text = f"⏰ {prefix} {rem_type}! Примерная сумма: {_format_amount(amount_rub, currency)}"
                    await bot.send_message(uid, text)

                    if next_date == today_str:
                        DB.mark_reminder_after_sent(int(r["id"]), periodicity)
                except Exception as inner_e:
                    print(f"[bot] reminder_loop send error: {inner_e}")
        except asyncio.CancelledError:
            return
        except Exception as e:
            print(f"[bot] reminder_loop error: {e}")
            await asyncio.sleep(60)


async def on_startup() -> None:
    print("[bot] startup: init_db")
    await asyncio.to_thread(DB.init_db)
    await asyncio.to_thread(FUEL_DB.init_db)
    global REMINDER_TASK
    if REMINDER_TASK is None:
        REMINDER_TASK = asyncio.create_task(reminder_loop())


async def on_shutdown() -> None:
    print("[bot] shutdown: close db")
    await asyncio.to_thread(DB.close)
    await asyncio.to_thread(FUEL_DB.close)
    global REMINDER_TASK
    if REMINDER_TASK is not None:
        REMINDER_TASK.cancel()
        REMINDER_TASK = None


async def main() -> None:
    bot = Bot(token=TOKEN)
    global BOT_REF
    BOT_REF = bot
    dp = Dispatcher()
    dp.include_router(router)

    dp.startup.register(on_startup)
    dp.shutdown.register(on_shutdown)

    # На случай перезапуска:
    await bot.delete_webhook(drop_pending_updates=True)

    print("[bot] polling started")
    await dp.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("[bot] stopped by user")
    except Exception as e:
        print(f"[bot] fatal error: {e}")

